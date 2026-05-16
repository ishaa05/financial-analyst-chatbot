"""
ingest.py  –  Parse every source document into typed chunks and store in ChromaDB.

Run once (or after adding new documents):
    python ingest.py

What it does:
  1. Walks `data/` and dispatches each file to the right parser.
  2. Produces chunks with rich metadata (doc_id, page, section, table_name, etc.).
  3. Embeds every chunk via Google's gemini-embedding-001 model.
  4. Upserts into a persistent ChromaDB collection so re-runs are idempotent.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import time
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Iterator

import chromadb
# from google import genai
# from google.genai import types
import openpyxl
import pandas as pd
import pdfplumber
from dotenv import load_dotenv
import xlrd
from sentence_transformers import SentenceTransformer
load_dotenv()
# client = genai.Client(
#     api_key=os.environ["GEMINI_API_KEY"],
#     http_options=types.HttpOptions(api_version="v1")
# )

# ─── Configuration ─────────────────────────────────────────────────────────────

DATA_DIR = Path("data")
CHROMA_DIR = Path("embeddings/chroma_db")
COLLECTION_NAME = "infosys_financials"
EMBED_MODEL = "BAAI/bge-large-en-v1.5"
model = SentenceTransformer(EMBED_MODEL)
EMBED_TASK = "RETRIEVAL_DOCUMENT"

# Chunking parameters
PDF_CHUNK_CHARS = 3000          # target characters per chunk
PDF_CHUNK_OVERLAP = 100        # overlap to preserve context across boundaries
MAX_EMBED_BATCH = 32          # Google API batch limit

# Document registry: filename stem → human label used in citations
DOC_REGISTRY = {
    "infosys-ar-25":              "Infosys Integrated Annual Report FY2024-25",
    "ifrs-usd-press-release_q1":  "Q1 FY26 Earnings Press Release (Jul 2025)",
    "ifrs-usd-press-release_q2":  "Q2 FY26 Earnings Press Release (Oct 2025)",
    "ifrs-usd-press-release_q3":  "Q3 FY26 Earnings Press Release (Jan 2026)",
    "ifrs-usd-press-release_q4":  "Q4 FY26 Earnings Press Release (Apr 2026)",
    "investor-sheet":             "Investor Multi-Year Financial Sheet",
    "500209":                     "Infosys BSE Stock Price History FY26",
}

# ─── Data model ────────────────────────────────────────────────────────────────

@dataclass
class Chunk:
    """One retrievable unit — a passage of text with its provenance."""
    chunk_id: str          # deterministic hash of (doc_id + content)
    doc_id: str            # key into DOC_REGISTRY
    doc_label: str         # human-readable citation label
    content: str           # the actual text
    chunk_type: str        # "text" | "table" | "spreadsheet_row" | "csv_row"
    page: int | None       # PDF page number (1-indexed)
    section: str           # best-guess section heading
    table_name: str        # table caption if chunk came from a table
    row_range: str         # "rows 5-20" for tabular chunks
    extra: dict            # any other metadata (ticker, quarter, sheet_name…)

    @classmethod
    def make_id(cls, doc_id: str, content: str, index: int = 0) -> str:
        return hashlib.sha256(f"{doc_id}:{index}:{content}".encode()).hexdigest()[:24]


# ─── PDF parser ────────────────────────────────────────────────────────────────

def _heading_candidates(page_text: str) -> list[str]:
    """Extract lines that look like section headings (short, no sentence-ending punctuation)."""
    candidates = []
    for line in page_text.split("\n"):
        line = line.strip()
        if 4 <= len(line) <= 80 and not line.endswith((".", ",", ";", ":")):
            candidates.append(line)
    return candidates


def _sliding_chunks(text: str, size: int, overlap: int) -> Iterator[str]:
    """Yield overlapping character-level windows over *text*."""
    start = 0
    while start < len(text):
        end = start + size
        # Try to break at a paragraph boundary
        boundary = text.rfind("\n\n", start, end)
        if boundary > start + size // 2:
            end = boundary
        yield text[start:end].strip()
        start = end - overlap
        if start >= len(text):
            break


def parse_pdf(path: Path) -> Iterator[Chunk]:
    """
    Extract text passages AND tables from a PDF.

    Strategy:
    - Text pages  → sliding-window chunks tagged with the last seen heading.
    - Table pages → one chunk per table, with the table serialised as markdown.
    """
    doc_id = path.stem
    label = DOC_REGISTRY.get(doc_id, path.name)
    current_section = "Introduction"

    with pdfplumber.open(path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            # ── Tables first ──────────────────────────────────────────────────
            for table_idx, table in enumerate(page.extract_tables()):
                if not table or len(table) < 2:
                    continue
                headers = [str(c or "").strip() for c in table[0]]
                rows = []
                for row in table[1:]:
                    rows.append({h: str(v or "").strip() for h, v in zip(headers, row)})

                # Convert table to clean markdown
                md_lines = ["| " + " | ".join(headers) + " |",
                             "| " + " | ".join(["---"] * len(headers)) + " |"]
                for row in rows:
                    md_lines.append("| " + " | ".join(row.values()) + " |")
                content = f"[TABLE — {current_section}]\n" + "\n".join(md_lines)

                yield Chunk(
                    chunk_id=Chunk.make_id(doc_id, content),
                    doc_id=doc_id,
                    doc_label=label,
                    content=content,
                    chunk_type="table",
                    page=page_num,
                    section=current_section,
                    table_name=f"Table {table_idx + 1} on page {page_num}",
                    row_range=f"rows 1-{len(rows)}",
                    extra={},
                )

            # ── Body text ─────────────────────────────────────────────────────
            text = page.extract_text() or ""
            if not text.strip():
                continue

            # Update running section heading
            headings = _heading_candidates(text)
            if headings:
                current_section = headings[0]

            for chunk_text in _sliding_chunks(text, PDF_CHUNK_CHARS, PDF_CHUNK_OVERLAP):
                if len(chunk_text) < 60:   # too short to be meaningful
                    continue
                content = f"[{current_section} — page {page_num}]\n{chunk_text}"
                yield Chunk(
                    chunk_id=Chunk.make_id(doc_id, content),
                    doc_id=doc_id,
                    doc_label=label,
                    content=content,
                    chunk_type="text",
                    page=page_num,
                    section=current_section,
                    table_name="",
                    row_range="",
                    extra={},
                )


def parse_excel(path: Path) -> Iterator[Chunk]:
    """
    Parse every sheet in the workbook.

    Each sheet is chunked in 25-row windows so that long P&L statements
    don't blow the context window but related rows stay together.
    """

    

    doc_id = path.stem
    label = DOC_REGISTRY.get(doc_id, path.name)

    ROW_WINDOW = 25
    ROW_OVERLAP = 5

    # ─── Handle old .xls files ─────────────────────────────
    if path.suffix.lower() == ".xls":

        wb = xlrd.open_workbook(path)

        for sheet in wb.sheets():

            rows = [sheet.row_values(i) for i in range(sheet.nrows)]

            if not rows:
                continue

            header_row = next((r for r in rows if any(c != "" for c in r)), None)

            if header_row is None:
                continue

            headers = [str(c).strip() for c in header_row]

            data_rows = rows[rows.index(header_row) + 1:]

            for start in range(0, len(data_rows), ROW_WINDOW - ROW_OVERLAP):

                window = data_rows[start:start + ROW_WINDOW]

                lines = [
                    "| " + " | ".join(headers) + " |",
                    "| " + " | ".join(["---"] * len(headers)) + " |"
                ]

                for row in window:
                    lines.append(
                        "| " + " | ".join(str(v or "") for v in row) + " |"
                    )

                content = (
                    f"[SPREADSHEET — sheet: {sheet.name}, "
                    f"rows {start + 1}–{start + len(window)}]\n"
                    + "\n".join(lines)
                )

                yield Chunk(
                    chunk_id=Chunk.make_id(doc_id, content),
                    doc_id=doc_id,
                    doc_label=label,
                    content=content,
                    chunk_type="spreadsheet_row",
                    page=None,
                    section=sheet.name,
                    table_name=sheet.name,
                    row_range=f"rows {start + 1}-{start + len(window)}",
                    extra={"sheet_name": sheet.name},
                )

        return

    # ─── Handle modern .xlsx files ────────────────────────
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:

        ws = wb[sheet_name]

        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        header_row = next((r for r in rows if any(c is not None for c in r)), None)

        if header_row is None:
            continue

        headers = [str(c or "").strip() for c in header_row]

        data_rows = rows[rows.index(header_row) + 1:]

        for start in range(0, len(data_rows), ROW_WINDOW - ROW_OVERLAP):

            window = data_rows[start:start + ROW_WINDOW]

            lines = [
                "| " + " | ".join(headers) + " |",
                "| " + " | ".join(["---"] * len(headers)) + " |"
            ]

            for row in window:
                lines.append(
                    "| " + " | ".join(str(v or "") for v in row) + " |"
                )

            content = (
                f"[SPREADSHEET — sheet: {sheet_name}, "
                f"rows {start + 1}–{start + len(window)}]\n"
                + "\n".join(lines)
            )

            yield Chunk(
                chunk_id=Chunk.make_id(doc_id, content),
                doc_id=doc_id,
                doc_label=label,
                content=content,
                chunk_type="spreadsheet_row",
                page=None,
                section=sheet_name,
                table_name=sheet_name,
                row_range=f"rows {start + 1}-{start + len(window)}",
                extra={"sheet_name": sheet_name},
            )

    wb.close()


# ─── CSV parser ────────────────────────────────────────────────────────────────

def parse_csv(path: Path) -> Iterator[Chunk]:
    """
    Parse a stock price CSV.
    Chunks by month so each chunk is semantically coherent (30-ish rows).
    Also yields a stats summary chunk for quick aggregate questions.
    """
    doc_id = path.stem
    label = DOC_REGISTRY.get(doc_id, path.name)

    df = pd.read_csv(path, parse_dates=True)

    # Try to detect and parse a date column
    date_col = next((c for c in df.columns if "date" in c.lower()), None)
    if date_col:
        df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
        df["_month"] = df[date_col].dt.to_period("M")
    else:
        df["_month"] = "all"

    # Per-month chunks
    for period, group in df.groupby("_month"):
        group_clean = group.drop(columns=["_month"])
        md = group_clean.to_markdown(index=False)
        content = f"[STOCK PRICES — {label} — {period}]\n{md}"
        yield Chunk(
            chunk_id=Chunk.make_id(doc_id, content),
            doc_id=doc_id,
            doc_label=label,
            content=content,
            chunk_type="csv_row",
            page=None,
            section=str(period),
            table_name="Stock prices",
            row_range=f"rows for {period}",
            extra={"period": str(period)},
        )

    # Aggregate stats chunk
    stats = df.drop(columns=["_month"], errors="ignore").describe().round(2)
    content = f"[STOCK PRICE SUMMARY STATISTICS — {label}]\n{stats.to_markdown()}"
    yield Chunk(
        chunk_id=Chunk.make_id(doc_id, content + "_stats"),
        doc_id=doc_id,
        doc_label=label,
        content=content,
        chunk_type="csv_row",
        page=None,
        section="Summary statistics",
        table_name="Stock price statistics",
        row_range="full year",
        extra={"is_summary": True},
    )


# ─── Embedding ─────────────────────────────────────────────────────────────────

def embed_batch(texts: list[str]) -> list[list[float]]:
    embeddings = model.encode(
        texts,
        show_progress_bar=False,
        convert_to_numpy=True
    )
    return embeddings.tolist()


def embed_in_batches(chunks: list[Chunk]) -> list[list[float]]:
    all_embeddings = []
    for i in range(0, len(chunks), MAX_EMBED_BATCH):
        batch = chunks[i: i + MAX_EMBED_BATCH]
        texts = [c.content for c in batch]
        print(f"  Embedding batch {i // MAX_EMBED_BATCH + 1} "
              f"({len(texts)} chunks)…")
        retries = 3
        while retries:
            try:
                all_embeddings.extend(embed_batch(texts))
                time.sleep(15)
                break
            except Exception as e:
                retries -= 1
                if retries == 0:
                    raise
                print(f"    Retry after error: {e}")
                time.sleep(60)
    return all_embeddings


# ─── ChromaDB upsert ───────────────────────────────────────────────────────────

def _metadata_for_chroma(chunk: Chunk) -> dict:
    metadata = asdict(chunk)
    metadata.pop("content")
    metadata["page"] = chunk.page if chunk.page is not None else -1
    metadata["extra_json"] = json.dumps(chunk.extra, ensure_ascii=True)
    metadata.pop("extra")
    return metadata


def upsert_to_chroma(chunks: list[Chunk], embeddings: list[list[float]]) -> None:
    chroma_client = chromadb.PersistentClient(path=str(CHROMA_DIR))  # renamed
    collection = chroma_client.get_or_create_collection(
        name=COLLECTION_NAME,
        metadata={"hnsw:space": "cosine"},
    )
    collection.upsert(
        ids=[c.chunk_id for c in chunks],
        embeddings=embeddings,
        documents=[c.content for c in chunks],
        metadatas=[_metadata_for_chroma(c) for c in chunks],
    )
    print(f"  ✓ Upserted {len(chunks)} chunks. "
          f"Collection now has {collection.count()} total.")


# ─── Orchestrator ──────────────────────────────────────────────────────────────

def run_ingestion() -> None:
    #genai.configure(api_key=os.environ["GEMINI_API_KEY"])

    all_chunks: list[Chunk] = []

    for path in sorted(DATA_DIR.iterdir()):
        if path.suffix.lower() == ".pdf":
            print(f"Parsing PDF: {path.name}")
            all_chunks.extend(parse_pdf(path))
        elif path.suffix.lower() in {".xlsx", ".xls"}:
            print(f"Parsing Excel: {path.name}")
            all_chunks.extend(parse_excel(path))
        elif path.suffix.lower() == ".csv":
            print(f"Parsing CSV: {path.name}")
            all_chunks.extend(parse_csv(path))

    # Deduplicate by chunk_id, keeping first occurrence
    seen = {}
    for i, chunk in enumerate(all_chunks):
        new_id = Chunk.make_id(chunk.doc_id, chunk.content, i)
        chunk.chunk_id = new_id
        seen[new_id] = chunk
    all_chunks = list(seen.values())
    print(f"\nTotal chunks after dedup: {len(all_chunks)}")
    print("Embedding…")
    embeddings = embed_in_batches(all_chunks)

    print("Writing to ChromaDB…")
    CHROMA_DIR.mkdir(parents=True, exist_ok=True)
    upsert_to_chroma(all_chunks, embeddings)

    print("\n✅ Ingestion complete.")


if __name__ == "__main__":
    run_ingestion()