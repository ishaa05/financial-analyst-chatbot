"""
formatters.py  –  Convert LLM answers into downloadable PDF reports and Excel files.

PDF generation strategy:
  - Parse the markdown answer into sections (headings, paragraphs, tables, lists).
  - Render each section with appropriate fpdf2 typography.
  - Append a formatted sources/citations appendix.

Excel generation strategy:
  - Ask a second LLM call to extract structured JSON data from the answer.
  - Write each data table to its own sheet with professional formatting.
  - Auto-size columns, add header row styling, freeze the header row.
"""

from __future__ import annotations

import io
import json
import os
import re
import textwrap
from datetime import datetime
from pathlib import Path
from groq import Groq
#import google.generativeai as genai
import openpyxl
from fpdf import FPDF
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter

from engine import EngineResponse


# ─── Shared helpers ────────────────────────────────────────────────────────────

def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _output_path(prefix: str, ext: str) -> Path:
    out = Path("outputs")
    out.mkdir(exist_ok=True)
    return out / f"{prefix}_{_timestamp()}.{ext}"


# ─────────────────────────────────────────────────────────────────────────────
# PDF GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

INFOSYS_BLUE = (0, 102, 204)    # Infosys brand blue
DARK_GRAY    = (40, 40, 40)
MID_GRAY     = (90, 90, 90)
LIGHT_GRAY   = (245, 245, 245)
WHITE        = (255, 255, 255)


class InfysysPDF(FPDF):
    """Custom FPDF subclass with Infosys-branded header and footer."""

    def __init__(self, title: str = "Financial Analysis Report"):
        super().__init__()
        self.report_title = title
        self.set_auto_page_break(auto=True, margin=20)
        self.set_margins(20, 25, 20)

    def header(self):
        # Blue bar
        self.set_fill_color(*INFOSYS_BLUE)
        self.rect(0, 0, 210, 14, "F")
        # Company name
        self.set_font("Helvetica", "B", 9)
        self.set_text_color(*WHITE)
        self.set_xy(10, 4)
        self.cell(0, 6, "INFOSYS LIMITED  |  Financial Intelligence Report", ln=False)
        # Report title (right-aligned)
        self.set_xy(10, 4)
        self.cell(0, 6, self.report_title, align="R", ln=False)
        self.ln(10)
        self.set_text_color(*DARK_GRAY)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(*MID_GRAY)
        self.cell(0, 10, f"Generated {datetime.now().strftime('%d %b %Y, %H:%M IST')}  |  Page {self.page_no()}", align="C")


def _md_to_pdf(pdf: InfysysPDF, markdown_text: str) -> None:
    """
    Render markdown into the PDF.
    Supports: # headings, ## subheadings, bullet lists, tables (|…|), paragraphs.
    """
    lines = markdown_text.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i].rstrip()

        # ── Heading 1 ─────────────────────────────────────────────────────────
        if line.startswith("# "):
            pdf.ln(4)
            pdf.set_font("Helvetica", "B", 16)
            pdf.set_text_color(*INFOSYS_BLUE)
            pdf.set_x(20)
            pdf.multi_cell(170, 9, _safe(line[2:]))
            # underline
            pdf.set_draw_color(*INFOSYS_BLUE)
            pdf.set_line_width(0.5)
            x = pdf.get_x()
            y = pdf.get_y()
            pdf.line(20, y, 190, y)
            pdf.ln(2)
            pdf.set_text_color(*DARK_GRAY)

        # ── Heading 2 ─────────────────────────────────────────────────────────
        elif line.startswith("## "):
            pdf.ln(3)
            pdf.set_font("Helvetica", "B", 13)
            pdf.set_text_color(*INFOSYS_BLUE)
            pdf.set_x(20)
            pdf.multi_cell(170, 8, _safe(line[3:]))

        # ── Heading 3 ─────────────────────────────────────────────────────────
        elif line.startswith("### "):
            pdf.ln(2)
            pdf.set_font("Helvetica", "B", 11)
            pdf.set_text_color(*MID_GRAY)
            pdf.set_x(20)
            pdf.multi_cell(0, 7, _safe(line[4:]))

        # ── Markdown table ────────────────────────────────────────────────────
        elif line.startswith("|") and i + 1 < len(lines) and lines[i + 1].startswith("|---"):
            # Collect all table rows
            table_lines = []
            while i < len(lines) and lines[i].startswith("|"):
                table_lines.append(lines[i])
                i += 1
            _render_table(pdf, table_lines)
            pdf.ln(3)
            continue   # i already advanced

        # ── Bullet / list item ────────────────────────────────────────────────
        elif re.match(r"^[-*•]\s", line):
            pdf.set_font("Helvetica", "", 10)
            text = re.sub(r"^\s*[-*•]\s+", "", line)
            text = _strip_inline_md(text)
            pdf.set_x(25)
            pdf.cell(5, 6, "-")
            pdf.set_x(30)
            pdf.set_x(20)
            pdf.multi_cell(160, 6, text)

        # ── Source citation line ([source:N] or ## Sources) ───────────────────
        elif line.startswith("## Sources") or line.startswith("- ["):
            if line.startswith("## Sources"):
                pdf.ln(4)
                pdf.set_font("Helvetica", "B", 11)
                pdf.set_text_color(*MID_GRAY)
                pdf.set_x(20)
                pdf.multi_cell(170, 7, "Sources")
                pdf.set_text_color(*DARK_GRAY)
            else:
                # Strip [source:N] markers for cleaner output
                clean = re.sub(r"\[source:\d+\]", "", line).strip()
                pdf.set_font("Helvetica", "I", 9)
                pdf.set_text_color(*MID_GRAY)
                pdf.set_x(25)
                pdf.multi_cell(155, 5, _safe(clean))
                pdf.set_text_color(*DARK_GRAY)

        # ── Separator ─────────────────────────────────────────────────────────
        elif re.match(r"^[-=]{3,}$", line):
            pdf.ln(2)
            pdf.set_draw_color(*MID_GRAY)
            pdf.line(20, pdf.get_y(), 190, pdf.get_y())
            pdf.ln(2)

        # ── Empty line ────────────────────────────────────────────────────────
        elif not line.strip():
            pdf.ln(3)

        # ── Normal paragraph ─────────────────────────────────────────────────
        else:
            pdf.set_font("Helvetica", "", 10)
            text = _strip_inline_md(line)
            # Remove [source:N] from body text
            text = re.sub(r"\[source:\d+\]", "", text).strip()
            if text:
                pdf.set_x(20)
                pdf.multi_cell(170, 6, text)

        i += 1


def _safe(text: str) -> str:
    """Sanitize text for FPDF latin-1 rendering."""
    text = text.replace("\u2026", "...").replace("\u2013", "-").replace("\u2014", "--")
    text = text.replace("\u2018", "'").replace("\u2019", "'")
    text = text.replace("\u201c", '"').replace("\u201d", '"')
    text = text.encode("latin-1", errors="ignore").decode("latin-1")
    return text.strip()


def _render_table(pdf: InfysysPDF, table_lines: list[str]) -> None:
    """Render a markdown table into the PDF as a formatted grid."""
    rows = []
    for line in table_lines:
        if re.match(r"^\|[-: |]+\|$", line):
            continue
        cells = [_safe(c.strip()) for c in line.strip("|").split("|")]
        rows.append(cells)

    if not rows:
        return

    headers = rows[0]
    data_rows = rows[1:]
    col_count = len(headers)

    if col_count == 0:
        return

    # Too many columns — fall back to plain text
    if col_count > 6:
        pdf.set_font("Helvetica", "I", 9)
        pdf.set_text_color(*MID_GRAY)
        pdf.set_x(20)
        pdf.multi_cell(0, 5, "[Wide table — download Excel for full data]")
        pdf.set_text_color(*DARK_GRAY)
        return

    available_w = 170
    col_w = max(available_w / col_count, 20)  # minimum 20mm per column
    max_chars = max(int(col_w / 1.8), 6)

    pdf.ln(2)
    # Header row
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(*INFOSYS_BLUE)
    pdf.set_text_color(*WHITE)
    for h in headers:
        pdf.cell(col_w, 7, _safe(h)[:max_chars], border=0, fill=True, align="C")
    pdf.ln()

    # Data rows
    pdf.set_font("Helvetica", "", 9)
    for row_idx, row in enumerate(data_rows):
        if row_idx % 2 == 0:
            pdf.set_fill_color(*LIGHT_GRAY)
        else:
            pdf.set_fill_color(*WHITE)
        pdf.set_text_color(*DARK_GRAY)
        padded = (row + [""] * col_count)[:col_count]
        for cell in padded:
            pdf.cell(col_w, 6, _safe(str(cell))[:max_chars], border=0, fill=True)
        pdf.ln()

    pdf.set_text_color(*DARK_GRAY)


def _strip_inline_md(text: str) -> str:
    """Remove inline markdown and sanitize non-latin characters for FPDF."""
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)
    text = re.sub(r"\*(.+?)\*",     r"\1", text)
    text = re.sub(r"`(.+?)`",       r"\1", text)
    text = re.sub(r"_(.+?)_",       r"\1", text)
    # Replace unicode punctuation with ASCII equivalents
    text = text.replace("\u2026", "...").replace("\u2013", "-").replace("\u2014", "--")
    text = text.replace("\u2018", "'").replace("\u2019", "'")
    text = text.replace("\u201c", '"').replace("\u201d", '"')
    # Drop any remaining non-latin1 characters
    text = text.encode("latin-1", errors="ignore").decode("latin-1")
    return text


def generate_pdf(response: EngineResponse, query: str) -> Path:
    """
    Build a branded PDF report from an EngineResponse.

    The report includes:
    - Title page with query and generation timestamp
    - Main answer rendered from markdown
    - Appendix: all retrieved source citations
    """
    # Derive a short report title from the query
    title_text = query[:60] + ("..." if len(query) > 60 else "")

    pdf = InfysysPDF(title=title_text)
    pdf.add_page()

   # ── Cover block ────────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 20)
    pdf.set_text_color(*INFOSYS_BLUE)
    pdf.ln(8)
    pdf.set_x(20)
    pdf.multi_cell(170, 12, "Infosys Financial Intelligence")
    pdf.set_font("Helvetica", "", 12)
    pdf.set_text_color(*MID_GRAY)
    safe_query = query.encode("latin-1", errors="ignore").decode("latin-1")
    pdf.set_x(20)
    pdf.multi_cell(170, 8, f"Query: {safe_query}")

    # ── Main answer ────────────────────────────────────────────────────────────
    _md_to_pdf(pdf, response.answer)

    # ── Sources appendix ──────────────────────────────────────────────────────
    if response.citations:
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 14)
        pdf.set_text_color(*INFOSYS_BLUE)
        pdf.cell(0, 10, "Appendix: Source References", ln=True)
        pdf.set_draw_color(*INFOSYS_BLUE)
        pdf.line(20, pdf.get_y(), 190, pdf.get_y())
        pdf.ln(4)

        for cite in response.citations:
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(*DARK_GRAY)
            page_str = f", page {cite['page']}" if cite.get("page") else ""
            section_str = f" - {_safe(str(cite.get('section', '')))}"  if cite.get("section") else ""
            cite_line = _safe(f"[{cite['id']}] {cite['label']}{page_str}{section_str}")
            pdf.set_x(20)
            pdf.multi_cell(170, 7, cite_line)
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(*MID_GRAY)
            for chunk in response.sources:
                if chunk.doc_label == cite["label"]:
                    raw = chunk.content.replace("\n", " ").replace("\r", " ")
                    raw = re.sub(r"\s+", " ", raw).strip()
                    snippet = _safe(raw[:400])
                    pdf.set_x(20)
                    pdf.multi_cell(170, 5, f'"{snippet}..."')
                    pdf.ln(2)
                    break

    out_path = _output_path("infosys_report", "pdf")
    pdf.output(str(out_path))
    return out_path


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

EXCEL_EXTRACTION_PROMPT = """You are a data extraction assistant.

Given this financial analyst answer (in Markdown), extract ALL numerical/tabular data
into a JSON structure that can be written to Excel.

Return ONLY valid JSON, no explanation, no markdown fences.
Format:
{{
  "sheets": [
    {{
      "name": "Sheet name (max 31 chars)",
      "description": "One-line description of what this sheet contains",
      "headers": ["Col1", "Col2", "Col3"],
      "rows": [
        ["val1", "val2", "val3"],
        ["val1", "val2", "val3"]
      ]
    }}
  ]
}}

Rules:
- Create one sheet per distinct data table or metric group in the answer.
- If a time series exists (quarters, years), put it in one sheet with periods as rows.
- Include units in the header names (e.g. "Revenue (USD Mn)").
- Every cell value must be a string or number — no nested objects.
- If no tabular data exists, create one sheet with a "Summary" header and the key facts.

Answer to extract from:
{answer}
"""


def _extract_excel_data(answer: str) -> list[dict]:
    """Ask Groq to extract structured tables from the markdown answer."""
    prompt = EXCEL_EXTRACTION_PROMPT.format(answer=answer)
    try:
        _client = Groq(api_key=os.environ["GROQ_API_KEY"])
        response = _client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_tokens=2048,
        )
        raw = response.choices[0].message.content.strip()
        raw = re.sub(r"^```(?:json)?", "", raw, flags=re.MULTILINE).strip()
        raw = re.sub(r"```$", "", raw, flags=re.MULTILINE).strip()
        # Find the first { and last } to extract just the JSON object
        start = raw.find("{")
        end = raw.rfind("}") + 1
        if start == -1 or end == 0:
            raise ValueError("No JSON object found in response")
        raw = raw[start:end]
        data = json.loads(raw)
        return data.get("sheets", [])
    except Exception as e:
        print(f"Excel extraction failed: {e}")
        # Fallback: parse markdown table from answer directly
        sheets = []
        lines = answer.split("\n")
        table_lines = [l for l in lines if l.strip().startswith("|")]
        if table_lines:
            headers = [c.strip() for c in table_lines[0].strip("|").split("|")]
            rows = []
            for line in table_lines[2:]:  # skip header and separator
                cells = [c.strip() for c in line.strip("|").split("|")]
                if cells:
                    rows.append(cells)
            sheets.append({
                "name": "Data",
                "description": "Extracted from answer",
                "headers": headers,
                "rows": rows,
            })
        return sheets


def _style_header_row(ws, header_row_num: int, col_count: int) -> None:
    """Apply Infosys-blue header styling to a row."""
    header_fill  = PatternFill(fill_type="solid", fgColor="0066CC")
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="FFFFFF")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, col_count + 1):
        cell = ws.cell(row=header_row_num, column=col)
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = header_align
        cell.border = border


def _style_data_rows(ws, start_row: int, end_row: int, col_count: int) -> None:
    """Alternate row shading for data rows."""
    fill_even = PatternFill(fill_type="solid", fgColor="EBF3FF")
    fill_odd  = PatternFill(fill_type="solid", fgColor="FFFFFF")
    data_font = Font(size=10)
    data_align = Alignment(vertical="center")

    for row_idx in range(start_row, end_row + 1):
        fill = fill_even if (row_idx - start_row) % 2 == 0 else fill_odd
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill  = fill
            cell.font  = data_font
            cell.alignment = data_align


def _auto_size_columns(ws) -> None:
    """Set column widths based on the longest value in each column."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 50)


def generate_excel(response: EngineResponse, query: str) -> Path:
    """
    Build a formatted Excel workbook from an EngineResponse.

    Steps:
    1. Ask Gemini to extract structured table data from the markdown answer.
    2. Write each table to its own sheet.
    3. Add an "About" sheet with the original query and source list.
    """
    sheets_data = _extract_excel_data(response.answer)

    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Data sheets ───────────────────────────────────────────────────────────
    for sheet_def in sheets_data:
        name = sheet_def.get("name", "Data")[:31]
        headers = sheet_def.get("headers", [])
        rows = sheet_def.get("rows", [])

        ws = wb.create_sheet(title=name)

        # Description row
        desc = sheet_def.get("description", "")
        if desc:
            ws.append([desc])
            ws["A1"].font = Font(italic=True, color="555555", size=9)
            ws.append([])   # blank row

        start_row = ws.max_row + 1

        # Header
        ws.append(headers)
        _style_header_row(ws, ws.max_row, len(headers))
        ws.row_dimensions[ws.max_row].height = 22

        # Data rows — try to coerce numbers
        data_start = ws.max_row + 1
        for row in rows:
            typed_row = []
            for val in row:
                try:
                    typed_row.append(float(str(val).replace(",", "")))
                except (ValueError, TypeError):
                    typed_row.append(val)
            ws.append(typed_row)

        _style_data_rows(ws, data_start, ws.max_row, len(headers))
        _auto_size_columns(ws)
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)   # freeze header

    # ── About sheet ───────────────────────────────────────────────────────────
    ws_about = wb.create_sheet(title="About", index=0)
    ws_about["A1"] = "Infosys Financial Intelligence – Data Export"
    ws_about["A1"].font = Font(bold=True, size=14, color="0066CC")

    ws_about["A3"] = "Query"
    ws_about["A3"].font = Font(bold=True)
    ws_about["B3"] = query

    ws_about["A4"] = "Generated"
    ws_about["A4"].font = Font(bold=True)
    ws_about["B4"] = datetime.now().strftime("%d %b %Y, %H:%M IST")

    ws_about["A6"] = "Sources used"
    ws_about["A6"].font = Font(bold=True)

    for row_offset, cite in enumerate(response.citations, start=7):
        page_str = f", page {cite['page']}" if cite.get("page") else ""
        ws_about.cell(row=row_offset, column=1, value=f"[{cite['id']}]")
        ws_about.cell(row=row_offset, column=2, value=f"{cite['label']}{page_str}")

    ws_about.column_dimensions["A"].width = 18
    ws_about.column_dimensions["B"].width = 60

    out_path = _output_path("infosys_data", "xlsx")
    wb.save(str(out_path))
    return out_path
